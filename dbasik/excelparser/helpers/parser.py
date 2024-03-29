import logging
import datetime
import numbers
import os
from enum import Enum, auto
from typing import Any, Dict, List, NamedTuple, Iterable

from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

from dbasik.datamap.models import Datamap, DatamapLine
from dbasik.register.models import Project
from dbasik.returns.models import Return, ReturnItem

SheetData = Dict[str, "WorkSheetFromDatamap"]

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)


class MissingSheetError(Exception):
    pass


class ParsedSpreadsheet:
    """
    A single spreadsheet whose data can be extracted using a Datamap upon
    calling the process() method. Data per sheet is then available via
    a processed_spreadsheet['sheet_name'] basis.
    """

    def __init__(
        self, template_path: str, project: Project, return_obj: Return, datamap: Datamap
    ) -> None:
        self.sheetnames: List[str]
        self.filename: str
        self.project_name = project.name
        self._template_path = template_path
        self._return_obj = return_obj
        self._datamap = datamap
        self._sheet_data: SheetData = {}
        self._get_sheets()
        self._get_filename()
        self._dml_sheets: List[str]
        self._dml_sheets_missing_from_spreadsheet: List[str]
        self._check_sheets_present()

        self._return_params = set(
            ["value_str", "value_int", "value_float", "value_date", "value_phone"]
        )

    def _map_to_keyword_param(self, cell_data: "CellData") -> str:
        _map = {
            CellValueType.STRING: "value_str",
            CellValueType.INTEGER: "value_int",
            CellValueType.FLOAT: "value_float",
            CellValueType.DATE: "value_date",
            CellValueType.PHONE: "value_phone",
        }
        # FIXME: too lazy to do anything but assign a string here
        return _map.get(cell_data.type, "value_str")

    def __getitem__(self, item):
        cls = type(self)
        if isinstance(item, numbers.Integral):
            msg = "{cls.__name__} indices must be strings"
            raise TypeError(msg.format(cls=cls))
        try:
            return self._sheet_data[item]
        except KeyError:
            msg = f"There is no sheet in the spreadsheet with title {item}."
            raise MissingSheetError(msg.format(item=item))

    def _get_filename(self):
        self.filename = os.path.split(self._template_path)[1]

    def _check_sheets_present(self) -> None:
        dmls = self._datamap.datamaplines.all()
        self._dml_sheets = list({dml.sheet for dml in dmls})
        _extra_sheet = list(set(self._dml_sheets) - set(self.sheetnames))
        if _extra_sheet:
            raise MissingSheetError(
                f"There is a worksheet in the spreadsheet not in the Datamap - {_extra_sheet[0]}"
            )

    def _process_sheets(self) -> None:
        wb: OpenpyxlWorkbook = load_workbook(self._template_path, data_only=True)
        logger.debug("Opening wb {}".format(wb))
        for ws in self.sheetnames:
            ws_from_dm = WorkSheetFromDatamap(
                openpyxl_worksheet=wb[ws], datamap=self._datamap
            )
            ws_from_dm._convert()
            self._sheet_data[ws] = ws_from_dm

    def process(self) -> None:
        """
        Convert a populated spreadsheet into a parseable data structure.
        :return: None
        :rtype: None
        """
        self._process_sheets()
        for sd in self._sheet_data.values():
            self._process_sheet_to_return(sd)

    def _process_sheet_to_return(self, sheet: "WorkSheetFromDatamap") -> None:
        sheet_name: str = sheet.title
        relevant_dmls: Iterable[DatamapLine] = self._datamap.datamaplines.filter(
            sheet=sheet_name
        )
        for dml in relevant_dmls:
            logger.debug(f"Processing {dml.key} in {dml.sheet}")
            _return_param = self._map_to_keyword_param(sheet[dml.key])
            logger.debug(f"_return_param: {_return_param}")
            _value_dict = {_return_param: sheet[dml.key].value}
            logger.debug(f"_value_dict: {_value_dict}")
            _other_params = self._return_params - set([_return_param])
            logger.debug(f"_other_params: {_other_params}")
            _combined_params = {k: None for k in list(_other_params)}
            _combined_params.update(_value_dict)
            logger.debug(f"_combined_params: {_combined_params}")
            ReturnItem.objects.create(
                parent=self.return_obj, datamapline=dml, **_combined_params
            )

    def _get_sheets(self) -> None:
        try:
            wb = load_workbook(self._template_path)
        except ImportError:
            raise
        self.sheetnames = wb.sheetnames

    @property
    def return_obj(self):
        return self._return_obj


class CellValueType(Enum):
    """
    Type classifiers for data parsed from a spreadsheet.
    """

    INTEGER = auto()
    STRING = auto()
    DATE = auto()
    FLOAT = auto()
    PHONE = auto()
    UNKNOWN = auto()


class CellData(NamedTuple):
    """
    Holds the data and useful metadata parsed from a spreadsheet.
    """

    key: str
    sheet: str
    value: Any
    source_cell: str
    type: CellValueType


class WorkSheetFromDatamap:
    """
    A dictionary-like object holding the data for a single spreadsheet sheet
    parsed using a Datamap object. Created by calling process() method on a
    ParsedSpreadsheet object.
    """

    def __init__(self, openpyxl_worksheet: OpenpyxlWorksheet, datamap: Datamap) -> None:
        self._data: Dict[str, CellData] = {}
        self._openpyxl_worksheet = openpyxl_worksheet
        self._datamap = datamap
        self._convert()
        self.title = self._openpyxl_worksheet.title

    def __getitem__(self, item):
        return self._data[item]

    def _convert(self) -> None:
        """
        Populates self._data dictionary with data from the spreadsheet.
        If type of data is not expected (i.e. not in the enum CellValueType)
        will still parse the data but classify it as CellValueType.UNKOWN
        for onward processing.
        :return: None
        :rtype: None
        """
        for _dml in self._datamap.datamaplines.filter(
            sheet__exact=self._openpyxl_worksheet.title
        ):
            _key = _dml.key

            # TODO the incorrect value is coming out here
            #  - it is coming out as date when using '\£#\m' format code

            _parsed_value = self._openpyxl_worksheet[_dml.cell_ref].value
            if isinstance(_parsed_value, datetime.datetime):
                _parsed_value = _parsed_value.date()
            _sheet_title = self._openpyxl_worksheet.title
            try:
                # TODO - in here we need to convert he phone number to a str
                _value = CellData(
                    _key,
                    _sheet_title,
                    _parsed_value,
                    _dml.cell_ref,
                    _detect_cell_type(_parsed_value, _dml),
                )
                self._data[_key] = _value
            except ValueError:
                _value = CellData(
                    _key,
                    _sheet_title,
                    _parsed_value,
                    _dml.cell_ref,
                    CellValueType.UNKNOWN,
                )
                _filter_phone_ints(_value)
                self._data[_key] = _value


def _filter_phone_ints(data: CellData) -> CellData:
    """
    Tests whether a CellData.value and CellData.type is an int and
    CellValueType.PHONE respectively. Converts the int to a str if true.
    """
    if isinstance(data.value, numbers.Integral) and data.type == CellValueType.PHONE:
        out = CellData(
            data.key, data.sheet, str(data.value), data.source_cell, data.type
        )
        return out
    else:
        return data


def _detect_cell_type(obj: Any, dml: DatamapLine) -> CellValueType:
    """
    Takes an object and maps its type to the CellValueType enum.
    Raises ValueError exception if the object is not an enum type
    useful for this process (int, str, float, etc).
    :param obj:
    :type obj: List[str]
    :return: CellValueType
    :rtype: None
    """
    if isinstance(obj, numbers.Integral):
        if dml.data_type == "Phone":
            return CellValueType.PHONE
        else:
            return CellValueType.INTEGER
    if isinstance(obj, str):
        return CellValueType.STRING
    if isinstance(obj, float):
        return CellValueType.FLOAT
    if isinstance(obj, (datetime.datetime, datetime.date)):
        return CellValueType.DATE
    else:
        raise ValueError("Cannot detect applicable type")
